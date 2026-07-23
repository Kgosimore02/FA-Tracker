"""Shared openpyxl helpers for all export views."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TEAL = "29B6D2"; NAVY = "1C2833"
_THIN   = Side(style="thin", color="CCCCCC")
BORDER  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CENTER  = Alignment(horizontal="center", vertical="center")
LEFT    = Alignment(horizontal="left",   vertical="center")

HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Arial", size=10)
TEAL_FONT = Font(name="Arial", bold=True, size=10, color=NAVY)

HDR_FILL  = PatternFill("solid", fgColor=NAVY)
TEAL_FILL = PatternFill("solid", fgColor=TEAL)

def title_row(ws, value: str, span: str):
    ws.merge_cells(span)
    c = ws[span.split(":")[0]]
    c.value = value
    c.font  = Font(name="Arial", bold=True, size=12, color=TEAL)
    c.alignment = CENTER

def hdr(ws, row, col, val, fill=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = HDR_FONT; c.fill = fill or HDR_FILL
    c.alignment = CENTER; c.border = BORDER

def cell(ws, row, col, val, fmt=None, align=LEFT):
    c = ws.cell(row=row, column=col, value=val)
    c.font = DATA_FONT; c.border = BORDER; c.alignment = align
    if fmt: c.number_format = fmt

def total_cell(ws, row, col, val, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = TEAL_FONT; c.fill = TEAL_FILL
    c.border = BORDER; c.alignment = CENTER
    if fmt: c.number_format = fmt

def col_widths(ws, pairs: list[tuple[str, int]]):
    for col, w in pairs:
        ws.column_dimensions[col].width = w
