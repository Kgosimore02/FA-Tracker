"""
export_monthly.py — Monthly Summary Export
Matches Weekly_Summary_Submissions.xlsx Section 3 (monthly summary table only).
One sheet per month. Pulls NOC + API budget from monthly_budgets table.
Region selected via dropdown.
"""
import io
from datetime import date
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from db import run_query

TEAL="FF29B6D2"; NAVY="FF1C2833"; WHITE="FFFFFFFF"
LGREY="FFF2F2F2"; MGREY="FFD9D9D9"; GREEN="FFC6EFCE"; RED="FFFFC7CE"
MN={1:"January",2:"February",3:"March",4:"April",5:"May",6:"June",
    7:"July",8:"August",9:"September",10:"October",11:"November",12:"December"}

def _f(bold=False,col="FF000000"): return Font(name="Arial",bold=bold,size=10,color=col)
def _p(h): return PatternFill("solid",fgColor=h)
def _b(): s=Side(style="thin"); return Border(left=s,right=s,top=s,bottom=s)
def _a(h="left"): return Alignment(horizontal=h,vertical="center")

def _s(ws,r,c,v,bold=False,fill=None,fmt=None,h="left",fc="FF000000"):
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=_f(bold,col=fc); cell.alignment=_a(h); cell.border=_b()
    if fill: cell.fill=_p(fill)
    if fmt:  cell.number_format=fmt
    return cell

def _build_monthly(ws, branches, prod_by_branch, budget_by_branch, year, mo, mname):
    n=len(branches); DR=range(3,3+n); TR=3+n
    ws.column_dimensions["A"].width=22
    for l in ["B","C","D","E","F","G","H"]: ws.column_dimensions[l].width=16

    # Title
    ws.merge_cells(f"A1:H1")
    c=ws["A1"]; c.value=f"Monthly Summary — {mname} {year}"
    c.font=_f(True,WHITE); c.fill=_p(NAVY); c.alignment=_a("center")

    # Headers
    hdrs=["Branch Name","NOC Budget","Monthly NOC","Over/Shortfall",
          "NOC % Attainment","API Budget (Monthly)","API Submission","API % Attainment"]
    for i,lbl in enumerate(hdrs,1):
        _s(ws,2,i,lbl,True,TEAL,h="center" if i>1 else "left",fc=WHITE)

    for ri,bname in zip(DR,branches):
        bg=budget_by_branch.get(bname,{}); p=prod_by_branch.get(bname,{})
        nb=bg.get("noc_budget",0) or 0
        ab=(bg.get("annual_premium_budget",0) or 0)/12
        noc=p.get("noc",0) or 0; api=p.get("api",0) or 0
        rf=LGREY if ri%2==0 else WHITE
        _s(ws,ri,1,bname,fill=rf)
        _s(ws,ri,2,nb,fill=rf,fmt="#,##0.00",h="right")
        _s(ws,ri,3,noc,fill=rf,fmt="#,##0",h="right")
        nos=noc-nb; _s(ws,ri,4,nos,fill=GREEN if nos>=0 else RED,fmt="#,##0.00",h="right")
        np=noc/nb if nb else 0; _s(ws,ri,5,np,fill=GREEN if np>=1 else RED,fmt="0.0%",h="right")
        _s(ws,ri,6,ab,fill=rf,fmt="#,##0.00",h="right")
        _s(ws,ri,7,api,fill=rf,fmt="#,##0.00",h="right")
        ap=api/ab if ab else 0; _s(ws,ri,8,ap,fill=GREEN if ap>=1 else RED,fmt="0.0%",h="right")

    # Totals
    _s(ws,TR,1,"Total",True,MGREY)
    for col in [2,3,6,7]:
        lc=get_column_letter(col)
        ws.cell(TR,col,f"=SUM({lc}{DR.start}:{lc}{DR.stop-1})").number_format="#,##0.00"
        ws.cell(TR,col).font=_f(True); ws.cell(TR,col).fill=_p(MGREY); ws.cell(TR,col).border=_b(); ws.cell(TR,col).alignment=_a("right")
    ws.cell(TR,4,f"=C{TR}-B{TR}").number_format="#,##0.00"
    ws.cell(TR,4).font=_f(True); ws.cell(TR,4).fill=_p(MGREY); ws.cell(TR,4).border=_b(); ws.cell(TR,4).alignment=_a("right")
    ws.cell(TR,5,f"=C{TR}/B{TR}").number_format="0.0%"
    ws.cell(TR,5).font=_f(True); ws.cell(TR,5).fill=_p(MGREY); ws.cell(TR,5).border=_b(); ws.cell(TR,5).alignment=_a("right")
    ws.cell(TR,8,f"=G{TR}/F{TR}").number_format="0.0%"
    ws.cell(TR,8).font=_f(True); ws.cell(TR,8).fill=_p(MGREY); ws.cell(TR,8).border=_b(); ws.cell(TR,8).alignment=_a("right")
    ws.freeze_panes="B3"


def _get_all_regions():
    return run_query("SELECT region_id,region_name FROM regions ORDER BY region_name")

def _get_branches(region_id):
    return [r["branch_name"] for r in run_query(
        "SELECT branch_name FROM branches WHERE region_id=:rid AND is_active=TRUE ORDER BY branch_name",
        {"rid":region_id})]

def _get_monthly_prod(region_id, year, month):
    rows = run_query("""
        SELECT b.branch_name, SUM(dp.noc_count) AS noc, SUM(dp.monthly_premium) AS api
        FROM daily_production dp
        JOIN field_agents fa ON fa.agent_id=dp.agent_id
        JOIN branches b ON b.branch_id=fa.branch_id
        WHERE b.region_id=:rid
          AND EXTRACT(YEAR FROM dp.production_date)=:yr
          AND EXTRACT(MONTH FROM dp.production_date)=:mo
        GROUP BY b.branch_name""",{"rid":region_id,"yr":year,"mo":month})
    return {r["branch_name"]:{"noc":float(r["noc"] or 0),"api":float(r["api"] or 0)} for r in rows}

def _get_budgets(region_id, year, month):
    rows = run_query("""
        SELECT b.branch_name, mb.noc_budget, mb.annual_premium_budget
        FROM monthly_budgets mb JOIN branches b ON b.branch_id=mb.branch_id
        WHERE b.region_id=:rid AND mb.budget_year=:yr AND mb.budget_month=:mo""",
        {"rid":region_id,"yr":year,"mo":month})
    return {r["branch_name"]:{"noc_budget":float(r["noc_budget"] or 0),
            "annual_premium_budget":float(r["annual_premium_budget"] or 0)} for r in rows}


def render(branch):
    st.title("📥 Monthly Summary Report")
    st.caption("Monthly NOC + API vs budget per branch, with over/shortfall and attainment %")

    regions=_get_all_regions()
    if not regions: st.error("No regions found."); return
    rmap={r["region_name"]:r["region_id"] for r in regions}
    own=run_query("SELECT region_id FROM branches WHERE branch_id=:bid",{"bid":branch["branch_id"]})
    own_rid=own[0]["region_id"] if own else regions[0]["region_id"]
    own_rname=next((r["region_name"] for r in regions if r["region_id"]==own_rid),regions[0]["region_name"])

    sel_region=st.selectbox("Region",list(rmap.keys()),
                            index=list(rmap.keys()).index(own_rname) if own_rname in rmap else 0)
    region_id=rmap[sel_region]

    yr=date.today().year
    year=st.selectbox("Year",list(range(yr,yr-3,-1)),index=0)
    sel_months=st.multiselect("Month(s)",list(MN.values()),default=[MN[date.today().month]])
    if not sel_months: st.warning("Select at least one month."); return

    if st.button("Generate",type="primary"):
        months=[k for k,v in MN.items() if v in sel_months]
        branches=_get_branches(region_id)
        with st.spinner("Building…"):
            wb=Workbook(); wb.remove(wb.active)
            for mo in sorted(months):
                mname=MN[mo]
                ws=wb.create_sheet(f"{mname[:3]} {year}")
                _build_monthly(ws,branches,_get_monthly_prod(region_id,year,mo),
                               _get_budgets(region_id,year,mo),year,mo,mname)
            buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        mstr="-".join(MN[m][:3] for m in sorted(months))
        st.download_button("⬇ Download Monthly Report",buf,
            f"Monthly_Summary_{sel_region.replace(' ','_')}_{year}_{mstr}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.success("Ready.")
