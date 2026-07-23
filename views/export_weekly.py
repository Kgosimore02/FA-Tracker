"""
export_weekly.py — Weekly NOC + API Submission Tracker Export
Matches Weekly_Summary_Submissions.xlsx exactly.

Data source priority:
  - weekly_noc_api table (backfilled from template for past months;
    entered by BMs going forward via the Weekly NOC Entry page)
  - Falls back to daily_production aggregated by week when weekly_noc_api
    has no data for a given branch/month

Layout per sheet: NOC side A-I | gap J | API side K-S
  Row 2 : section titles
  Row 3 : headers
  Rows 4…: branch data
  +0: Total, +1: Weekly Targets (15/20/30/35%), +2: Deficit, +3: % Attainment
  Row 21+: Section 2 — weekly targets breakdown per branch
  Row 34+: Section 3 — monthly summary
"""
import io
from datetime import date
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from db import run_query

TEAL="FF29B6D2"; NAVY="FF1C2833"; WHITE="FFFFFFFF"
LGREY="FFF2F2F2"; MGREY="FFD9D9D9"; DGREY="FF595959"
GREEN="FFC6EFCE"; RED="FFFFC7CE"
MN={1:"January",2:"February",3:"March",4:"April",5:"May",6:"June",
    7:"July",8:"August",9:"September",10:"October",11:"November",12:"December"}

# NOC cols A-I (1-9), gap J (10), API cols K-S (11-19)
NC={"br":1,"w1":2,"w2":3,"w3":4,"w4":5,"tot":6,"os":7,"bud":8,"pct":9}
AC={"br":11,"w1":12,"w2":13,"w3":14,"w4":15,"tot":16,"os":17,"bud":18,"pct":19}

def _f(bold=False,col="FF000000"): return Font(name="Arial",bold=bold,size=10,color=col)
def _p(h): return PatternFill("solid",fgColor=h)
def _b(): s=Side(style="thin"); return Border(left=s,right=s,top=s,bottom=s)
def _a(h="left"): return Alignment(horizontal=h,vertical="center")

def _s(ws,r,c,v,bold=False,fill=None,fmt=None,h="left",fc="FF000000"):
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=_f(bold,col=fc); cell.alignment=_a(h); cell.border=_b()
    if fill: cell.fill=_p(fill)
    if fmt: cell.number_format=fmt
    return cell

def _styled(ws,r,c,v,fill,bold=False,fmt=None,h="right",fc="FF000000"):
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=_f(bold,col=fc); cell.fill=_p(fill); cell.border=_b(); cell.alignment=_a(h)
    if fmt: cell.number_format=fmt
    return cell


def _build(ws, branches, data, budgets, year, mo, mname):
    n=len(branches); DR=range(4,4+n); TR=4+n; WR=TR+1; DEFR=TR+2; PCTER=TR+3
    S2H=21; S2D=range(22,22+n); S2T=22+n
    S3H=S2T+3; S3D=range(S3H+1,S3H+1+n); S3T=S3H+1+n

    # col widths
    ws.column_dimensions["A"].width=22; ws.column_dimensions["K"].width=22
    ws.column_dimensions["J"].width=2
    for c in range(2,10): ws.column_dimensions[get_column_letter(c)].width=13
    for c in range(12,20): ws.column_dimensions[get_column_letter(c)].width=14

    # Row 1 title
    ws.merge_cells("A1:S1")
    c=ws["A1"]; c.value=f"{mname} {year} — Weekly Submission Tracker"
    c.font=Font(name="Arial",bold=True,size=12,color=WHITE); c.fill=_p(NAVY); c.alignment=_a("center")

    # Row 2 section titles
    ws.merge_cells("A2:I2")
    c=ws["A2"]; c.value="Weekly NOC Submission tracker"
    c.font=_f(True,col=WHITE); c.fill=_p(TEAL); c.alignment=_a("center")
    ws.merge_cells("K2:S2")
    c=ws["K2"]; c.value="Weekly API Submission tracker"
    c.font=_f(True,col=WHITE); c.fill=_p(TEAL); c.alignment=_a("center")

    # Row 3 headers
    for i,lbl in enumerate(["Branch Name","Week 1","Week 2","Week 3","Week 4",
                              "Total","Over/Shortfall","NOC Budget","% Attainment"],1):
        c=ws.cell(3,i,lbl); c.font=_f(True,col=WHITE); c.fill=_p(NAVY)
        c.border=_b(); c.alignment=_a("left" if i==1 else "center")
    for i,lbl in enumerate(["Branch Name","Week 1","Week 2","Week 3","Week 4",
                              "Total","Over/Shortfall","API Budget","% Attainment"],11):
        c=ws.cell(3,i,lbl); c.font=_f(True,col=WHITE); c.fill=_p(NAVY)
        c.border=_b(); c.alignment=_a("left" if i==11 else "center")

    # Data rows
    for ri,bname in zip(DR,branches):
        bd=data.get(bname,{}); bg=budgets.get(bname,{})
        nb=bg.get("noc_budget",0) or 0
        ab=(bg.get("annual_premium_budget",0) or 0)
        rf=LGREY if ri%2==0 else WHITE

        _s(ws,ri,NC["br"],bname,fill=rf)
        nwv=[bd.get(w,{}).get("noc",0) or 0 for w in [1,2,3,4]]
        for col,v in zip([NC["w1"],NC["w2"],NC["w3"],NC["w4"]],nwv):
            _s(ws,ri,col,v,fill=rf,fmt="#,##0",h="right")
        nt=sum(nwv)
        _s(ws,ri,NC["tot"],nt,True,fill=rf,fmt="#,##0",h="right")
        nos=nt-nb; _styled(ws,ri,NC["os"],nos,GREEN if nos>=0 else RED,fmt="#,##0.00")
        _s(ws,ri,NC["bud"],nb,fill=rf,fmt="#,##0.00",h="right")
        np=nt/nb if nb else 0; _styled(ws,ri,NC["pct"],np,GREEN if np>=1 else RED,fmt="0.0%")

        _s(ws,ri,AC["br"],bname,fill=rf)
        awv=[bd.get(w,{}).get("api",0) or 0 for w in [1,2,3,4]]
        for col,v in zip([AC["w1"],AC["w2"],AC["w3"],AC["w4"]],awv):
            _s(ws,ri,col,v,fill=rf,fmt="#,##0.00",h="right")
        at=sum(awv)
        _s(ws,ri,AC["tot"],at,True,fill=rf,fmt="#,##0.00",h="right")
        aos=at-ab; _styled(ws,ri,AC["os"],aos,GREEN if aos>=0 else RED,fmt="#,##0.00")
        _s(ws,ri,AC["bud"],ab,fill=rf,fmt="#,##0.00",h="right")
        ap=at/ab if ab else 0; _styled(ws,ri,AC["pct"],ap,GREEN if ap>=1 else RED,fmt="0.0%")

    # Total row
    ds,de=DR.start,DR.stop-1
    _s(ws,TR,NC["br"],"Total",True,fill=MGREY)
    for col in [NC["w1"],NC["w2"],NC["w3"],NC["w4"],NC["tot"],NC["bud"]]:
        lc=get_column_letter(col)
        _styled(ws,TR,col,f"=SUM({lc}{ds}:{lc}{de})",MGREY,True,fmt="#,##0.00")
    fl=get_column_letter(NC["tot"]); hl=get_column_letter(NC["bud"])
    _styled(ws,TR,NC["os"],f"={fl}{TR}-{hl}{TR}",MGREY,True,fmt="#,##0.00")
    _styled(ws,TR,NC["pct"],f"={fl}{TR}/{hl}{TR}",MGREY,True,fmt="0.0%")

    _s(ws,TR,AC["br"],"Total",True,fill=MGREY)
    for col in [AC["w1"],AC["w2"],AC["w3"],AC["w4"],AC["tot"],AC["bud"]]:
        lc=get_column_letter(col)
        _styled(ws,TR,col,f"=SUM({lc}{ds}:{lc}{de})",MGREY,True,fmt="#,##0.00")
    pl=get_column_letter(AC["tot"]); rl=get_column_letter(AC["bud"])
    _styled(ws,TR,AC["os"],f"={pl}{TR}-{rl}{TR}",MGREY,True,fmt="#,##0.00")
    _styled(ws,TR,AC["pct"],f"={pl}{TR}/{rl}{TR}",MGREY,True,fmt="0.0%")

    # Weekly Targets row (15/20/30/35% of budget total)
    def _nv(ws,r,c,v,fmt=None):
        _styled(ws,r,c,v,NAVY,True,fmt,fc=WHITE)
    _s(ws,WR,NC["br"],"Weekly Targets",True,fill=NAVY,fc=WHITE)
    for col,pct in zip([NC["w1"],NC["w2"],NC["w3"],NC["w4"]],[.15,.20,.30,.35]):
        _nv(ws,WR,col,f"={get_column_letter(NC['tot'])}{TR}*{pct}","#,##0.00")
    _nv(ws,WR,NC["tot"],f"={get_column_letter(NC['bud'])}{TR}","#,##0.00")

    _s(ws,WR,AC["br"],"Weekly Targets",True,fill=NAVY,fc=WHITE)
    for col,pct in zip([AC["w1"],AC["w2"],AC["w3"],AC["w4"]],[.15,.20,.30,.35]):
        _nv(ws,WR,col,f"={get_column_letter(AC['tot'])}{TR}*{pct}","#,##0.00")
    _nv(ws,WR,AC["tot"],f"={get_column_letter(AC['bud'])}{TR}","#,##0.00")

    # Deficit row
    def _dv(ws,r,c,v,fmt=None):
        _styled(ws,r,c,v,DGREY,True,fmt,fc=WHITE)
    _s(ws,DEFR,NC["br"],"Deficit",True,fill=DGREY,fc=WHITE)
    for col in [NC["w1"],NC["w2"],NC["w3"],NC["w4"],NC["tot"]]:
        lc=get_column_letter(col); _dv(ws,DEFR,col,f"={lc}{TR}-{lc}{WR}","#,##0.00")
    _s(ws,DEFR,AC["br"],"Deficit",True,fill=DGREY,fc=WHITE)
    for col in [AC["w1"],AC["w2"],AC["w3"],AC["w4"],AC["tot"]]:
        lc=get_column_letter(col); _dv(ws,DEFR,col,f"={lc}{TR}-{lc}{WR}","#,##0.00")

    # % Attainment row
    def _tv(ws,r,c,v,fmt=None):
        _styled(ws,r,c,v,TEAL,True,fmt,fc=WHITE)
    _s(ws,PCTER,NC["br"],"% Attainment",True,fill=TEAL,fc=WHITE)
    for col in [NC["w1"],NC["w2"],NC["w3"],NC["w4"],NC["tot"]]:
        lc=get_column_letter(col); _tv(ws,PCTER,col,f"={lc}{TR}/{lc}{WR}","0.0%")
    _s(ws,PCTER,AC["br"],"% Attainment",True,fill=TEAL,fc=WHITE)
    for col in [AC["w1"],AC["w2"],AC["w3"],AC["w4"],AC["tot"]]:
        lc=get_column_letter(col); _tv(ws,PCTER,col,f"={lc}{TR}/{lc}{WR}","0.0%")

    # Section 2: Weekly Targets Breakdown
    ws.merge_cells(f"A{S2H-1}:S{S2H-1}")
    c=ws.cell(S2H-1,1,"Weekly Targets Breakdown — Pro-rated Budget per Branch")
    c.font=_f(True,col=WHITE); c.fill=_p(NAVY); c.alignment=_a("center")
    for i,lbl in enumerate(["Branch","Wk1 (15%)","Wk2 (20%)","Wk3 (30%)","Wk4 (35%)",f"{mname[:3]} Target"],1):
        _s(ws,S2H,i,lbl,True,TEAL,h="center" if i>1 else "left",fc=WHITE)
    for i,lbl in enumerate(["Branch","Wk1 (15%)","Wk2 (20%)","Wk3 (30%)","Wk4 (35%)",f"{mname[:3]} Target"],11):
        _s(ws,S2H,i,lbl,True,TEAL,h="center" if i>11 else "left",fc=WHITE)

    for ri,bname in zip(S2D,branches):
        bg=budgets.get(bname,{}); nb=bg.get("noc_budget",0) or 0
        ab=(bg.get("annual_premium_budget",0) or 0)/12
        rf=LGREY if ri%2==0 else WHITE
        _s(ws,ri,1,bname,fill=rf)
        for col,pct in zip([2,3,4,5],[.15,.20,.30,.35]):
            _s(ws,ri,col,nb*pct,fill=rf,fmt="#,##0.00",h="right")
        _s(ws,ri,6,nb,True,fill=rf,fmt="#,##0.00",h="right")
        _s(ws,ri,11,bname,fill=rf)
        for col,pct in zip([12,13,14,15],[.15,.20,.30,.35]):
            _s(ws,ri,col,ab*pct,fill=rf,fmt="#,##0.00",h="right")
        _s(ws,ri,16,ab,True,fill=rf,fmt="#,##0.00",h="right")

    _s(ws,S2T,1,"Total",True,MGREY)
    for col in [2,3,4,5,6]:
        lc=get_column_letter(col)
        _styled(ws,S2T,col,f"=SUM({lc}{S2D.start}:{lc}{S2D.stop-1})",MGREY,True,fmt="#,##0.00")
    _s(ws,S2T,11,"Total",True,MGREY)
    for col in [12,13,14,15,16]:
        lc=get_column_letter(col)
        _styled(ws,S2T,col,f"=SUM({lc}{S2D.start}:{lc}{S2D.stop-1})",MGREY,True,fmt="#,##0.00")

    # Section 3: Monthly Summary
    ws.merge_cells(f"A{S3H-1}:H{S3H-1}")
    c=ws.cell(S3H-1,1,f"Monthly Summary — {mname} {year}")
    c.font=_f(True,col=WHITE); c.fill=_p(NAVY); c.alignment=_a("center")
    for i,lbl in enumerate(["Branch Name","NOC Budget","Monthly Submission",
                              "Over/Shortfall","NOC % Attainment","API % Attainment","Variance"],1):
        _s(ws,S3H,i,lbl,True,TEAL,h="center" if i>1 else "left",fc=WHITE)

    nb_c=get_column_letter(NC["bud"]); nt_c=get_column_letter(NC["tot"])
    np_c=get_column_letter(NC["pct"]); ap_c=get_column_letter(AC["pct"])
    for ri,bname,src in zip(S3D,branches,DR):
        rf=LGREY if ri%2==0 else WHITE
        _s(ws,ri,1,bname,fill=rf)
        for col,ref,fmt in [(2,f"={nb_c}{src}","#,##0.00"),
                             (3,f"={nt_c}{src}","#,##0"),
                             (5,f"={np_c}{src}","0.0%"),
                             (6,f"={ap_c}{src}","0.0%")]:
            cell=ws.cell(ri,col,ref); cell.number_format=fmt
            cell.fill=_p(rf); cell.border=_b(); cell.font=_f(); cell.alignment=_a("right")
        ws.cell(ri,4,f"=C{ri}-B{ri}").number_format="#,##0.00"
        ws.cell(ri,4).fill=_p(rf); ws.cell(ri,4).border=_b(); ws.cell(ri,4).font=_f(); ws.cell(ri,4).alignment=_a("right")
        ws.cell(ri,7,f"=C{ri}-B{ri}").number_format="#,##0"
        ws.cell(ri,7).fill=_p(rf); ws.cell(ri,7).border=_b(); ws.cell(ri,7).font=_f(); ws.cell(ri,7).alignment=_a("right")

    _s(ws,S3T,1,"Total",True,MGREY)
    for col in [2,3]:
        lc=get_column_letter(col)
        _styled(ws,S3T,col,f"=SUM({lc}{S3D.start}:{lc}{S3D.stop-1})",MGREY,True,fmt="#,##0.00")
    _styled(ws,S3T,4,f"=C{S3T}-B{S3T}",MGREY,True,fmt="#,##0.00")
    _styled(ws,S3T,5,f"=C{S3T}/B{S3T}",MGREY,True,fmt="0.0%")
    _styled(ws,S3T,6,f"={ap_c}{TR}",MGREY,True,fmt="0.0%")
    ws.freeze_panes="B4"


# ── DB queries ────────────────────────────────────────────────────────────────

def _all_regions():
    return run_query("SELECT region_id,region_name FROM regions ORDER BY region_name")

def _branches(region_id):
    return [r["branch_name"] for r in run_query(
        "SELECT branch_name FROM branches WHERE region_id=:rid AND is_active=TRUE ORDER BY branch_name",
        {"rid":region_id})]

def _get_data(region_id, year, month):
    """
    Returns {branch_name: {week: {noc, api}}}.
    Prefers weekly_noc_api; falls back to daily_production bucketed by week.
    """
    # Try weekly_noc_api first
    rows = run_query("""
        SELECT b.branch_name, w.week_number,
               w.noc_count AS noc, w.monthly_premium AS api
        FROM weekly_noc_api w
        JOIN branches b ON b.branch_id=w.branch_id
        WHERE b.region_id=:rid AND w.year=:yr AND w.month=:mo
        ORDER BY b.branch_name, w.week_number""",
        {"rid":region_id,"yr":year,"mo":month})

    if rows:
        out={}
        for r in rows:
            b=r["branch_name"]
            if b not in out: out[b]={}
            out[b][r["week_number"]]={"noc":int(r["noc"] or 0),"api":float(r["api"] or 0)}
        return out

    # Fall back to daily_production bucketed by Mon-Sun week
    # week_number = ISO week position within the month (1-4)
    rows = run_query("""
        SELECT b.branch_name,
               LEAST(4, (EXTRACT(DAY FROM dp.production_date)::int - 1) / 7 + 1)::int AS wk,
               SUM(dp.noc_count) AS noc, SUM(dp.monthly_premium) AS api
        FROM daily_production dp
        JOIN field_agents fa ON fa.agent_id=dp.agent_id
        JOIN branches b ON b.branch_id=fa.branch_id
        WHERE b.region_id=:rid
          AND EXTRACT(YEAR FROM dp.production_date)=:yr
          AND EXTRACT(MONTH FROM dp.production_date)=:mo
        GROUP BY b.branch_name, wk ORDER BY b.branch_name, wk""",
        {"rid":region_id,"yr":year,"mo":month})

    out={}
    for r in rows:
        b=r["branch_name"]
        if b not in out: out[b]={}
        out[b][r["wk"]]={"noc":int(r["noc"] or 0),"api":float(r["api"] or 0)}
    return out

def _budgets(region_id, year, month):
    rows=run_query("""
        SELECT b.branch_name, mb.noc_budget, mb.annual_premium_budget
        FROM monthly_budgets mb JOIN branches b ON b.branch_id=mb.branch_id
        WHERE b.region_id=:rid AND mb.budget_year=:yr AND mb.budget_month=:mo""",
        {"rid":region_id,"yr":year,"mo":month})
    return {r["branch_name"]:{"noc_budget":float(r["noc_budget"] or 0),
            "annual_premium_budget":float(r["annual_premium_budget"] or 0)} for r in rows}

def _own_region(branch):
    rows=run_query("SELECT region_id FROM branches WHERE branch_id=:bid",{"bid":branch["branch_id"]})
    return rows[0]["region_id"] if rows else None


# ── Streamlit render ──────────────────────────────────────────────────────────

def render(branch):
    st.title("📥 Weekly NOC/API Report")
    st.caption("Matches Weekly_Summary_Submissions.xlsx — NOC + API with budgets, weekly targets, deficit")

    regions=_all_regions()
    if not regions: st.error("No regions found."); return
    rmap={r["region_name"]:r["region_id"] for r in regions}
    own_rid=_own_region(branch)
    own_name=next((r["region_name"] for r in regions if r["region_id"]==own_rid), list(rmap.keys())[0])

    sel=st.selectbox("Region", list(rmap.keys()),
                     index=list(rmap.keys()).index(own_name) if own_name in rmap else 0)
    region_id=rmap[sel]

    yr=date.today().year
    year=st.selectbox("Year", list(range(yr,yr-3,-1)), index=0)
    sel_months=st.multiselect("Month(s)", list(MN.values()), default=[MN[date.today().month]])
    if not sel_months: st.warning("Select at least one month."); return

    if st.button("Generate", type="primary"):
        months=[k for k,v in MN.items() if v in sel_months]
        branches=_branches(region_id)
        with st.spinner("Building…"):
            wb=Workbook(); wb.remove(wb.active)
            for mo in sorted(months):
                ws=wb.create_sheet(f"{MN[mo][:3]} {year}")
                _build(ws, branches, _get_data(region_id,year,mo),
                       _budgets(region_id,year,mo), year, mo, MN[mo])
            buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        mstr="-".join(MN[m][:3] for m in sorted(months))
        st.download_button("⬇ Download Weekly NOC/API Report", buf,
            f"Weekly_NOC_API_{sel.replace(' ','_')}_{year}_{mstr}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.success("Ready.")
