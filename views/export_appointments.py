"""
export_appointments.py — Weekly Appointment Tracker Export
Matches Weekly_appointment_tracker.xlsx exactly.
Region selected via dropdown. Data from weekly_appointments table.
"""
import io
from datetime import date
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from db import run_query

TEAL="FF29B6D2"; NAVY="FF1C2833"; WHITE="FFFFFFFF"
LGREY="FFF2F2F2"; MGREY="FFD9D9D9"; GREEN="FFC6EFCE"; RED="FFFFC7CE"; YEL="FFFFFFCC"
MN={1:"January",2:"February",3:"March",4:"April",5:"May",6:"June",
    7:"July",8:"August",9:"September",10:"October",11:"November",12:"December"}

def _f(bold=False,col="FF000000"): return Font(name="Arial",bold=bold,size=10,color=col)
def _p(h): return PatternFill("solid",fgColor=h)
def _b(): s=Side(style="thin"); return Border(left=s,right=s,top=s,bottom=s)
def _a(h="left"): return Alignment(horizontal=h,vertical="center")
def _s(ws,r,c,v,bold=False,fill=None,fmt=None,h="left",fc="FF000000"):
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=_f(bold,col=fc); cell.alignment=_a(h); cell.border=_b()
    if fill: cell.fill=_p(fill)
    if fmt:  cell.number_format=fmt
    return cell

def _build_appt(ws, branches, appt_by_branch, fa_counts, year, mo, mname):
    n=len(branches); DR=range(4,4+n); TR=4+n
    ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=11
    ws.column_dimensions["C"].width=20
    for l in ["D","E","G","H","J","K","M","N"]: ws.column_dimensions[l].width=13
    for l in ["F","I","L"]: ws.column_dimensions[l].width=2
    ws.column_dimensions["O"].width=10

    # Row 1 title
    ws.merge_cells("A1:O1")
    c=ws["A1"]; c.value="Weekly appoinment tracker"
    c.font=_f(True,WHITE); c.fill=_p(NAVY); c.alignment=_a("center")

    # Row 2 week spans
    for sc,ec,lbl in [(4,5,"Week 1"),(7,8,"Week 2"),(10,11,"Week 3"),(13,14,"Week 4")]:
        ws.merge_cells(start_row=2,start_column=sc,end_row=2,end_column=ec)
        c=ws.cell(2,sc,lbl); c.font=_f(True,WHITE); c.fill=_p(TEAL); c.alignment=_a("center")
    c=ws.cell(2,15,"Total "); c.font=_f(True,WHITE); c.fill=_p(NAVY); c.alignment=_a("center")

    # Row 3 headers
    hmap={1:"Branch Name",2:"Number FA",3:"Target number of Appointments",
          4:"Attained",5:"Success rate %",7:"Attained",8:"Success rate %",
          10:"Attained",11:"Success rate %",13:"Attained",14:"Success rate %"}
    for col,lbl in hmap.items():
        c=ws.cell(3,col,lbl); c.font=_f(True,WHITE); c.fill=_p(NAVY); c.border=_b()
        c.alignment=_a("left" if col==1 else "center")

    # Data rows
    for ri,bname in zip(DR,branches):
        ap=appt_by_branch.get(bname,{}); fa=fa_counts.get(bname,0) or 0
        rf=LGREY if ri%2==0 else WHITE
        _s(ws,ri,1,bname,fill=rf)
        _s(ws,ri,2,fa,fill=rf,fmt="#,##0",h="right")
        ws.cell(ri,3,f"=B{ri}*3").number_format="#,##0"
        ws.cell(ri,3).fill=_p(rf); ws.cell(ri,3).border=_b(); ws.cell(ri,3).font=_f(); ws.cell(ri,3).alignment=_a("right")
        for wk,ac,sc in [(1,4,5),(2,7,8),(3,10,11),(4,13,14)]:
            att=ap.get(wk,0) or 0; al=get_column_letter(ac)
            _s(ws,ri,ac,att,fill=rf,fmt="#,##0",h="right")
            sr=att/(fa*3) if fa else 0
            sf=GREEN if sr>=1 else YEL if sr>=0.7 else (rf if att==0 else RED)
            ws.cell(ri,sc,f"={al}{ri}/C{ri}").number_format="0.0%"
            ws.cell(ri,sc).fill=_p(sf); ws.cell(ri,sc).border=_b(); ws.cell(ri,sc).font=_f(); ws.cell(ri,sc).alignment=_a("right")
        ws.cell(ri,15,f"=SUM(G{ri},J{ri},M{ri})").number_format="#,##0"
        ws.cell(ri,15).font=_f(True); ws.cell(ri,15).fill=_p(rf); ws.cell(ri,15).border=_b(); ws.cell(ri,15).alignment=_a("right")

    # TOTAl row
    ds,de=DR.start,DR.stop-1
    _s(ws,TR,1,"TOTAl",True,MGREY,fc=NAVY)
    for col in [2,3]:
        lc=get_column_letter(col)
        ws.cell(TR,col,f"=SUM({lc}{ds}:{lc}{de})").number_format="#,##0"
        ws.cell(TR,col).font=_f(True); ws.cell(TR,col).fill=_p(MGREY); ws.cell(TR,col).border=_b(); ws.cell(TR,col).alignment=_a("right")
    for ac,sc in [(4,5),(7,8),(10,11),(13,14)]:
        al=get_column_letter(ac)
        ws.cell(TR,ac,f"=SUM({al}{ds}:{al}{de})").number_format="#,##0"
        ws.cell(TR,ac).font=_f(True); ws.cell(TR,ac).fill=_p(MGREY); ws.cell(TR,ac).border=_b(); ws.cell(TR,ac).alignment=_a("right")
        ws.cell(TR,sc,f"={al}{TR}/C{TR}").number_format="0.0%"
        ws.cell(TR,sc).font=_f(True); ws.cell(TR,sc).fill=_p(MGREY); ws.cell(TR,sc).border=_b(); ws.cell(TR,sc).alignment=_a("right")
    ws.cell(TR,15,f"=SUM(G{TR},J{TR},M{TR})").number_format="#,##0"
    ws.cell(TR,15).font=_f(True); ws.cell(TR,15).fill=_p(MGREY); ws.cell(TR,15).border=_b(); ws.cell(TR,15).alignment=_a("right")
    ws.freeze_panes="B4"


def _get_all_regions():
    return run_query("SELECT region_id,region_name FROM regions ORDER BY region_name")

def _get_branches(region_id):
    return [r["branch_name"] for r in run_query(
        "SELECT branch_name FROM branches WHERE region_id=:rid AND is_active=TRUE ORDER BY branch_name",
        {"rid":int(region_id)})]

def _get_appointments(region_id, year, month):
    from db import run_query_live
    rows=run_query_live("""
        SELECT b.branch_name,wa.week_number,wa.appointments_attained
        FROM weekly_appointments wa JOIN branches b ON b.branch_id=wa.branch_id
        WHERE b.region_id=:rid AND wa.year=:yr AND wa.month=:mo""",
        {"rid":int(region_id),"yr":int(year),"mo":int(month)})
    out={}
    for r in rows:
        b=r["branch_name"]
        if b not in out: out[b]={}
        out[b][r["week_number"]]=r["appointments_attained"] or 0
    return out

def _get_fa_counts(region_id, year, month):
    from db import run_query_live
    rows=run_query_live("""
        SELECT b.branch_name,
               COALESCE(
                   MAX(wa.fa_count),
                   COUNT(*) FILTER (WHERE fa.is_active=TRUE)
               ) AS fa_count
        FROM branches b
        LEFT JOIN field_agents fa ON fa.branch_id=b.branch_id
        LEFT JOIN weekly_appointments wa
            ON wa.branch_id=b.branch_id
            AND wa.year=:yr AND wa.month=:mo
        WHERE b.region_id=:rid AND b.is_active=TRUE
        GROUP BY b.branch_name""",
        {"rid":int(region_id),"yr":int(year),"mo":int(month)})
    return {r["branch_name"]:int(r["fa_count"] or 0) for r in rows}


def render(branch):
    st.title("📥 Weekly Appointments Export")
    st.caption("Matches Weekly_appointment_tracker.xlsx — appointments attained vs target per branch")

    regions=_get_all_regions()
    if not regions: st.error("No regions found."); return
    rmap={r["region_name"]:r["region_id"] for r in regions}
    own=run_query("SELECT region_id FROM branches WHERE branch_id=:bid",{"bid":branch["branch_id"]})
    own_rid=own[0]["region_id"] if own else regions[0]["region_id"]
    own_rname=next((r["region_name"] for r in regions if r["region_id"]==own_rid),regions[0]["region_name"])

    sel_region=st.selectbox("Region",list(rmap.keys()),
                            index=list(rmap.keys()).index(own_rname) if own_rname in rmap else 0)
    region_id=rmap[sel_region]

    yr=date.today().year
    year=st.selectbox("Year",list(range(yr,yr-3,-1)),index=0)
    sel_months=st.multiselect("Month(s)",list(MN.values()),default=[MN[date.today().month]])
    if not sel_months: st.warning("Select at least one month."); return

    
    if st.button("Generate",type="primary"):
        st.cache_data.clear()
        months=[k for k,v in MN.items() if v in sel_months]
        branches=_get_branches(region_id)
        with st.spinner("Building…"):
            wb=Workbook(); wb.remove(wb.active)
            for mo in sorted(months):
                mname=MN[mo]
                fa_counts=_get_fa_counts(region_id, year, mo)
                ws=wb.create_sheet(f"{mname[:3]} {year}")
                _build_appt(ws,branches,_get_appointments(region_id,year,mo),fa_counts,year,mo,mname)
            buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        mstr="-".join(MN[m][:3] for m in sorted(months))
        st.download_button("⬇ Download Appointments",buf,
            f"Weekly_Appointments_{sel_region.replace(' ','_')}_{year}_{mstr}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.success("Ready.")